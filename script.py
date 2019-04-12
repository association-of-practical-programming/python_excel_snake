import openpyxl;
from openpyxl.styles import Color, PatternFill, Font, Border

wb = openpyxl.load_workbook('Book1.xlsx')

ws = wb.active

ws['A1'] = 4

redFill = PatternFill(start_color='FFFFFFFF',
                   end_color='FFFF0000',
                   fill_type='solid')


ws['A1'].fill = redFill

for sheet in wb:
    print(sheet.title)


wb.save('Book2.xlsx')